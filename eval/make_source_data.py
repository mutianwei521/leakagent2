# -*- coding: utf-8 -*-
"""Build the Nature Water Source Data / Supplementary Data workbooks.

Every cell is written programmatically from the committed artifacts under
artifacts/ (plus, for the zoning workbook, the vendored data/ inputs the
figures themselves read). Nothing is hand-entered. Chinese field values in
the City D / City H records (DMA names, leak-type labels) are mapped to
stable anonymized codes before they reach a submission workbook; the private
code-to-name mapping is written next to the workbooks and is NOT part of the
submission set.

Run:  python eval/make_source_data.py
Out:  build_sd/*.xlsx  +  build_sd/SD_MANIFEST.md  +  build_sd/PRIVATE_*.json
"""
import os
import re
import sys
import json
import pickle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ART = os.path.join(ROOT, "artifacts")
OUT = os.path.join(ROOT, "build_sd")
os.makedirs(OUT, exist_ok=True)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

BODY = Font(name="Arial", size=10)
HEAD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=12, bold=True)
NOTE = Font(name="Arial", size=9, italic=True)
FILL = PatternFill("solid", fgColor="DCE6F1")

LEAK_TYPE_EN = {"明漏": "visible (surface) leak",
                "暗漏": "hidden (buried) leak"}

CJK = re.compile(r"[　-鿿豈-﫿]")


def L(name):
    return json.load(open(os.path.join(ART, name), encoding="utf-8"))


def fmt(v):
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (list, tuple)):
        return "; ".join(str(x) for x in v)
    return v


def sheet(wb, title, header, rows, notes=(), widths=None):
    ws = wb.create_sheet(title=title[:31])
    ws.append([str(h) for h in header])
    for c in ws[1]:
        c.font = HEAD
        c.fill = FILL
    for r in rows:
        ws.append([fmt(v) for v in r])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = BODY
    if notes:
        ws.append([])
        for n in notes:
            ws.append([n])
            ws.cell(ws.max_row, 1).font = NOTE
    for j, h in enumerate(header, start=1):
        w = (widths or {}).get(j, max(10, min(38, len(str(h)) + 4)))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    return ws


def readme(wb, title, desc_lines, prov_rows):
    ws = wb.create_sheet(title="README", index=0)
    ws["A1"] = title
    ws["A1"].font = TITLE
    r = 3
    for line in desc_lines:
        ws.cell(r, 1, line).font = BODY
        r += 1
    r += 1
    hdr = ["Sheet", "Display item", "Source artifact (committed)", "JSON key(s)", "Producing run script"]
    for j, h in enumerate(hdr, start=1):
        ws.cell(r, j, h).font = HEAD
        ws.cell(r, j).fill = FILL
    for row in prov_rows:
        r += 1
        for j, v in enumerate(row, start=1):
            ws.cell(r, j, v).font = BODY
    for j, w in ((1, 30), (2, 26), (3, 40), (4, 52), (5, 40)):
        ws.column_dimensions[get_column_letter(j)].width = w
    r += 2
    ws.cell(r, 1, "All values are written by eval/make_source_data.py directly from the named artifact;"
                  " no cell is hand-entered.").font = NOTE
    return ws


def save(wb, fname):
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"].max_column == 1:
        del wb["Sheet"]
    path = os.path.join(OUT, fname)
    wb.save(path)
    return path


def seeds_of(prov):
    s = prov.get("seeds")
    return [str(x) for x in s] if s else ["s1", "s2", "s3", "s4", "s5"]


def metric_rows(block, seed_labels, keys=None):
    rows = []
    for k in (keys or block.keys()):
        m = block[k]
        rows.append([k, m["mean"], m["std"]] + list(m["values"]))
    header = ["metric (fraction)", "mean", "sd (ddof=1)"] + ["seed " + s for s in seed_labels]
    return header, rows


# ======================================================================
# workbook 1: Source Data Fig. 4  (EXA7 evidence figure)
# ======================================================================
def build_fig4():
    rf = L("results_full.json")
    bf = L("results_baselines_fair.json")
    st = L("results_stats.json")
    sl = seeds_of(rf["provenance"])
    wb = Workbook()

    cc = rf["risk_coverage_sigma0.05_seed42"]
    ck = list(cc[0].keys())
    sheet(wb, "Fig4a_contract_curve", ck, [[p[k] for k in ck] for p in cc],
          notes=["Acceptance-threshold sweep for the full goal contract, seed 42, sigma = 0.05 m."])

    ec = rf["risk_coverage_existence_only_sigma0.05_seed42"]
    ek = list(ec[0].keys())
    sheet(wb, "Fig4a_existence_curve", ek, [[p[k] for k in ek] for p in ec],
          notes=["Existence-threshold-only baseline sweep, same events and noise."])

    m = rf["multiseed_sigma0.05"]["forced_retrieval_top1"]
    sheet(wb, "Fig4a_retrieval_reference", ["seed", "forced retrieval top-1 (fraction)"],
          [[s, v] for s, v in zip(sl, m["values"])],
          notes=["Red dashed reference line in panel a is the mean of this column: "
                 "mean = {} , sd (ddof=1) = {}.".format(m["mean"], m["std"])])

    dc = rf["differential_confusion_sigma0.05_seed42"]
    sheet(wb, "Fig4b_confusion", ["true class \\ predicted"] + list(dc["classes"]),
          [[cls] + list(row) for cls, row in zip(dc["classes"], dc["matrix"])],
          notes=["Raw event counts, seed 42; panel colours are row-normalised rates."])

    hdr = ["method", "mean (fraction)", "sd (ddof=1)"] + ["seed " + s for s in sl]
    rows = [[k, v["mean"], v["std"]] + list(v["values"]) for k, v in bf["forced_top1"].items()]
    sheet(wb, "Fig4c_baselines", hdr, rows,
          notes=["Forced top-1 zone accuracy on identical held-out test leaks; 5 seeds."])

    ps = st["per_severity_leak_acc"]
    rows = [[k, ps[k]["n"], ps[k]["acc"]] + list(ps[k]["values"]) for k in ps]
    sheet(wb, "Fig4d_severity", ["injected discharge", "n events", "pooled accuracy (fraction)"]
          + ["seed " + s for s in sl], rows)

    ns = rf["noise_sweep_multiseed"]
    rows = []
    for sig in ns:
        for met in ("leak_acc", "coverage_at_100prec"):
            b = ns[sig][met]
            rows.append([sig, met, b["mean"], b["std"]] + list(b["values"]))
    sheet(wb, "Fig4e_noise_sweep", ["sigma (m)", "metric (fraction)", "mean", "sd (ddof=1)"]
          + ["seed " + s for s in sl], rows)

    readme(wb, "Source Data for Fig. 4 (EXA7 benchmark: selective decision quality,"
               " differential diagnosis, forced baselines)",
           ["One sheet per panel. Fractions are plotted as percentages in the figure.",
            "Seeds: " + ", ".join(sl) + "; sensor noise sigma = 0.05 m unless a sheet states otherwise."],
           [["Fig4a_contract_curve", "Fig. 4a (blue)", "artifacts/results_full.json",
             "risk_coverage_sigma0.05_seed42", "eval/run_experiments.py via reproduce_all.py"],
            ["Fig4a_existence_curve", "Fig. 4a (grey)", "artifacts/results_full.json",
             "risk_coverage_existence_only_sigma0.05_seed42", "eval/run_experiments.py"],
            ["Fig4a_retrieval_reference", "Fig. 4a (red dashed)", "artifacts/results_full.json",
             "multiseed_sigma0.05.forced_retrieval_top1", "eval/run_experiments.py"],
            ["Fig4b_confusion", "Fig. 4b", "artifacts/results_full.json",
             "differential_confusion_sigma0.05_seed42", "eval/run_experiments.py"],
            ["Fig4c_baselines", "Fig. 4c", "artifacts/results_baselines_fair.json",
             "forced_top1", "eval/baselines_fair.py"],
            ["Fig4d_severity", "Fig. 4d", "artifacts/results_stats.json",
             "per_severity_leak_acc", "eval/run_full_stats.py"],
            ["Fig4e_noise_sweep", "Fig. 4e", "artifacts/results_full.json",
             "noise_sweep_multiseed.{leak_acc, coverage_at_100prec}", "eval/run_experiments.py"]])
    return save(wb, "Source_Data_Fig4.xlsx")


# ======================================================================
# City D anonymization helpers
# ======================================================================
def dma_map_for(rows, prefix):
    seen = {}
    for r in rows:
        d = r.get("dma")
        if d is not None and d not in seen:
            seen[d] = "DMA {}-{}".format(prefix, len(seen) + 1)
    return seen


def anon_row(r, dmap, fields):
    out = []
    for f in fields:
        v = r.get(f)
        if f == "dma":
            v = dmap.get(v, v)
        elif f == "type":
            v = LEAK_TYPE_EN.get(v, v)
        out.append(fmt(v))
    return out


# ======================================================================
# workbook 2: Source Data Fig. 5  (L-Town + City D register)
# ======================================================================
def build_fig5():
    lt = L("results_ltown.json")
    dg = L("results_city_d.json")
    sv = L("results_city_d_severity.json")
    wb = Workbook()

    fr = lt["risk_coverage"]["frontier"]
    fk = list(fr[0].keys())
    sheet(wb, "Fig5a_frontier", fk, [[p[k] for k in fk] for p in fr],
          notes=["Existence-threshold frontier over the 33 labelled L-Town leaks."])
    sm = lt["summary"]
    sheet(wb, "Fig5a_operating_point", ["quantity", "value"], [[k, fmt(v)] for k, v in sm.items()],
          notes=["The starred operating point in panel a is (coverage, decision_precision_on_acted)."])

    ev_fields = ["pipe", "type", "true_zones", "max_abs_dp_m", "existence", "margin",
                 "outcome", "pred_zone", "top1_zone", "correct"]
    sheet(wb, "Fig5b_events", ev_fields,
          [[fmt(r.get(f)) for f in ev_fields] for r in lt["rows"]],
          notes=["Panel b plots existence (x) against margin (y) for all 33 leaks."])
    cvs = lt["contract_vs_scalar_threshold"]
    sheet(wb, "Fig5b_contract_vs_scalar", ["quantity", "value"],
          [[k, fmt(v)] for k, v in cvs.items()],
          notes=["The margin floor drawn in panel b is the contract constant delta = 0.10"
                 " (Methods, Eq. 4, predicate G3); it is a preset threshold, not a fitted value."])

    dmap = dma_map_for(dg["rows"], "D")
    reg_fields = ["id", "type", "dma", "diameter_mm", "rate_Ls", "severity_tpd",
                  "max_abs_dp_clean_m", "max_abs_dp_m", "outcome", "true_zones",
                  "accepted_zone", "pred_zone", "top1_zone", "retrieval_top1",
                  "twinfit_top1", "correct", "existence", "margin", "top1_mahal", "best_alt"]
    hdr = ["event", "leak type", "DMA (coded)", "diameter (mm)", "rate (L/s)",
           "severity (t/day)", "clean peak |dp| (m)", "observed peak |dp| (m)", "outcome",
           "true zone(s)", "accepted zone", "predicted zone", "top-1 zone",
           "retrieval top-1 zone", "twin-fit top-1 zone", "zone correct", "existence",
           "margin", "top-1 Mahalanobis per dof", "best alternative posterior"]
    sheet(wb, "Fig5c_register_events", hdr,
          [anon_row(r, dmap, reg_fields) for r in dg["rows"]],
          notes=["All 194 audited 2025 work orders. DMA names are coded for utility"
                 " confidentiality; leak-type labels are translated from the register.",
                 "Panel c plots rate (L/s) against clean peak |dp| (m), coloured by outcome."])
    det = dg["detectability"]
    sheet(wb, "Fig5c_thresholds", ["quantity", "value"],
          [[k, fmt(v)] for k, v in det.items()] + [["sensor_std_m", dg["summary"]["sensor_std_m"]]],
          notes=["Red dashed line: actionable_dp_threshold_m; amber dotted line: sensor_std_m."])

    sw = sv["sweep"]
    sk = list(sw[0].keys())
    sheet(wb, "Fig5d_severity_sweep", sk, [[p[k] for k in sk] for p in sw],
          notes=["Controlled single-leak sweep at {} seeded nodes.".format(sv.get("n_test_nodes")),
                 "The register-rate histogram in panel d is built from the rate (L/s) column"
                 " of sheet Fig5c_register_events."])

    readme(wb, "Source Data for Fig. 5 (L-Town benchmark and City D field register)",
           ["One sheet per panel. Fractions are plotted as percentages in the figure.",
            "City D DMA names are replaced by stable codes (DMA D-1, D-2, ...) for utility"
            " confidentiality; every numeric value is unchanged."],
           [["Fig5a_frontier", "Fig. 5a", "artifacts/results_ltown.json", "risk_coverage.frontier",
             "eval/run_ltown.py"],
            ["Fig5a_operating_point", "Fig. 5a (star)", "artifacts/results_ltown.json", "summary",
             "eval/run_ltown.py"],
            ["Fig5b_events", "Fig. 5b", "artifacts/results_ltown.json", "rows", "eval/run_ltown.py"],
            ["Fig5b_contract_vs_scalar", "Fig. 5b", "artifacts/results_ltown.json",
             "contract_vs_scalar_threshold", "eval/run_ltown.py"],
            ["Fig5c_register_events", "Fig. 5c", "artifacts/results_city_d.json", "rows",
             "eval/run_city_d.py"],
            ["Fig5c_thresholds", "Fig. 5c (floors)", "artifacts/results_city_d.json",
             "detectability, summary.sensor_std_m", "eval/run_city_d.py"],
            ["Fig5d_severity_sweep", "Fig. 5d", "artifacts/results_city_d_severity.json", "sweep",
             "eval/run_city_d_severity.py"]])
    return save(wb, "Source_Data_Fig5.xlsx"), dmap


# ======================================================================
# workbook 3: Source Data Table 1  (the identical measurement battery)
# ======================================================================
def build_table1():
    rf = L("results_full.json")
    ky = L("results_ky4.json")
    hs = L("results_city_h_standard.json")
    ds = L("results_city_d_standard.json")
    ub = L("results_unified_battery.json")
    lt = L("results_ltown.json")
    dg = L("results_city_d.json")
    ft = L("results_city_d_flowtier.json")
    dc = L("results_city_d_controls.json")
    sc = L("results_city_d_std_controls.json")
    wb = Workbook()

    h, r = metric_rows(rf["multiseed_sigma0.05"], seeds_of(rf["provenance"]))
    sheet(wb, "EXA7_per_seed", h, r)
    for name, art, fname in (("KY4_per_seed", ky, "results_ky4.json"),
                             ("CityH_per_seed", hs, "results_city_h_standard.json"),
                             ("CityD_per_seed", ds, "results_city_d_standard.json")):
        h, r = metric_rows(art["multiseed_sigma0.05"], seeds_of(art["provenance"]))
        pv = art["provenance"]
        notes = ["Provenance: " + "; ".join("{} = {}".format(k, fmt(v)) for k, v in pv.items()
                 if k in ("network", "n_junctions", "n_pipes", "partition_k", "n_sensors",
                          "seeds", "n_scenarios_per_seed", "noise_m"))]
        sheet(wb, name, h, r, notes=notes)

    rows = []
    for net, label in (("exa7", "EXA7"), ("ky4", "KY4"), ("city_h", "City H"), ("city_d", "City D")):
        p = ub[net]["pooled"]
        c = ub[net]["no_leak_controls"]
        g = ub[net]["per_seed_gates"]
        rows.append([label, p["n_events"], p["n_acted"], p["coverage"], p["contract_precision"],
                     p["contract_correct"], p["scalar_existence_precision_matched"],
                     p["scalar_correct"], "{}/{}".format(c["dispatched"], c["n"]),
                     c["false_alarm_rate"], "; ".join("{}:{}".format(k, v) for k, v in g.items())])
    sheet(wb, "Battery_pooled", ["network", "n events", "n acted", "coverage (fraction)",
                                 "contract precision", "contract correct",
                                 "scalar precision (matched coverage)", "scalar correct",
                                 "no-leak controls dispatched", "false alarm rate",
                                 "per-seed operating gates"], rows,
          notes=["Table 1 rows 'Coverage', 'Precision on acted', 'No-leak controls' and"
                 " 'Contract vs scalar' for the four in-silico columns come from this sheet."])

    sheet(wb, "LTown_column", ["quantity", "value"],
          [[k, fmt(v)] for k, v in lt["summary"].items()]
          + [["contract_vs_scalar." + k, fmt(v)] for k, v in lt["contract_vs_scalar_threshold"].items()],
          notes=["L-Town no-leak controls are n/a: no labelled leak-free window exists"
                 " (artifacts/results_ltown_timeline.json)."])

    surv = [(k, v) for k, v in ft["results"].items()
            if v["survey_dispatched"] == v["zone_correct"] and v["survey_dispatched"] > 0]
    surv_rows = [[k] + [fmt(v[c]) for c in ("sigma_f_Ls", "u", "n_events", "survey_dispatched",
                                            "zone_correct", "coverage", "precision_on_dispatched",
                                            "controls_false_alarm")] for k, v in ft["results"].items()]
    sheet(wb, "CityD_register_column", ["quantity", "value"],
          [[k, fmt(v)] for k, v in dg["summary"].items()]
          + [["controls.false_dispatches." + k, fmt(v)] for k, v in dc["false_dispatches"].items()]
          + [["std_controls.false_dispatches." + k, fmt(v)] for k, v in sc["false_dispatches"].items()],
          notes=["Excavation tier: summary.n_acted of n_leaks_evaluated, precision"
                 " summary.decision_precision_on_acted.",
                 "Survey tier settings are in sheet CityD_survey_tier."])
    sheet(wb, "CityD_survey_tier", ["setting", "sigma_f (L/s)", "u", "n events", "survey dispatched",
                                    "zone correct", "coverage (fraction)", "precision on dispatched",
                                    "controls false alarm"], surv_rows,
          notes=["Table 1 quotes the district-survey tier at full district precision;"
                 " the qualifying settings (dispatched = correct) are: "
                 + ", ".join(k for k, _ in surv) + "."])

    readme(wb, "Source Data for Table 1 (the identical measurement battery on all five networks)",
           ["Per-seed sheets carry every metric behind the in-silico columns;"
            " Battery_pooled carries the shared-gate battery cells;"
            " the two real-tier columns have their own sheets.",
            "Fractions correspond to the percentages printed in Table 1."],
           [["EXA7_per_seed", "Table 1 col. EXA7", "artifacts/results_full.json",
             "multiseed_sigma0.05", "eval/run_experiments.py"],
            ["KY4_per_seed", "Table 1 col. KY4", "artifacts/results_ky4.json",
             "multiseed_sigma0.05", "eval/run_ky4.py"],
            ["CityH_per_seed", "Table 1 col. City H", "artifacts/results_city_h_standard.json",
             "multiseed_sigma0.05", "eval/run_standard_net.py city_h"],
            ["CityD_per_seed", "Table 1 col. City D", "artifacts/results_city_d_standard.json",
             "multiseed_sigma0.05", "eval/run_standard_net.py city_d"],
            ["Battery_pooled", "Table 1 battery rows", "artifacts/results_unified_battery.json",
             "{exa7,ky4,city_h,city_d}.{pooled,no_leak_controls,per_seed_gates}",
             "eval/run_unified_battery.py"],
            ["LTown_column", "Table 1 col. L-Town", "artifacts/results_ltown.json",
             "summary, contract_vs_scalar_threshold", "eval/run_ltown.py"],
            ["CityD_register_column", "Table 1 col. City D register",
             "artifacts/results_city_d.json + results_city_d_controls.json + results_city_d_std_controls.json",
             "summary, false_dispatches", "eval/run_city_d.py, eval/run_city_d_controls.py"],
            ["CityD_survey_tier", "Table 1 col. City D register (survey)",
             "artifacts/results_city_d_flowtier.json", "results", "eval/run_city_d_flow_tier.py"]])
    return save(wb, "Source_Data_Table1.xlsx")


# ======================================================================
# workbook 4: Source Data Fig. S2 (multiseed robustness)
# ======================================================================
def build_figs2():
    rf = L("results_full.json")
    wb = Workbook()
    keys = ["forced_retrieval_top1", "leak_partition_acc_no_active",
            "leak_partition_acc_with_active", "decision_precision_at_maxcov"]
    h, r = metric_rows(rf["multiseed_sigma0.05"], seeds_of(rf["provenance"]), keys)
    sheet(wb, "FigS2_multiseed", h, r)
    readme(wb, "Source Data for Supplementary Fig. S2 (seed-to-seed robustness, EXA7)",
           ["The four bars of Fig. S2 with per-seed replicate dots."],
           [["FigS2_multiseed", "Fig. S2", "artifacts/results_full.json",
             "multiseed_sigma0.05 (four of the six metrics)", "eval/run_experiments.py"]])
    return save(wb, "Source_Data_FigS2.xlsx")


# ======================================================================
# workbook 5: Source Data Figs. 3, S1, S3 (zoning and network maps)
# ======================================================================
def build_fig3_maps():
    wb = Workbook()

    def part_at(parts, K):
        return parts[K] if K in parts else parts[str(K)]

    with open(os.path.join(ROOT, "data", "exa7", "partitions.pkl"), "rb") as fh:
        parts = pickle.load(fh)
    n2c = part_at(parts, 15)["node_to_community"]
    sensors = json.load(open(os.path.join(ROOT, "data", "exa7", "sensor_fingerprints.json"),
                             encoding="utf-8"))["sensor_nodes"]
    sheet(wb, "EXA7_node_to_zone", ["node id", "zone (K=15)", "is sensor"],
          [[n, z, "yes" if n in set(sensors) else ""] for n, z in sorted(n2c.items())],
          notes=["Full node-to-zone assignment for the vendored EXA7 network"
                 " (data/exa7/partitions.pkl, K = 15) and the 30 greedy-placed sensors."])

    rows = []
    for net, K in (("exa7", 15), ("ky4", 25), ("ltown", 25), ("city_h", 25), ("city_d", 15)):
        try:
            with open(os.path.join(ROOT, "data", net, "partitions.pkl"), "rb") as fh:
                p = pickle.load(fh)
            nc = part_at(p, K)["node_to_community"]
            sn = json.load(open(os.path.join(ROOT, "data", net, "sensor_fingerprints.json"),
                                encoding="utf-8"))
            ns = len(sn.get("sensor_nodes") or sn.get("sensors") or [])
            sizes = {}
            for z in nc.values():
                sizes[z] = sizes.get(z, 0) + 1
            label = {"exa7": "EXA7", "ky4": "KY4", "ltown": "L-Town",
                     "city_h": "City H", "city_d": "City D"}[net]
            for z in sorted(sizes):
                rows.append([label, K, z, sizes[z], ns])
        except FileNotFoundError:
            continue
    sheet(wb, "Zone_sizes_all_networks", ["network", "K", "zone", "n nodes in zone",
                                          "n sensors (network)"], rows,
          notes=["Zone-size distributions behind the maps of Fig. 3 and Figs. S1/S3.",
                 "Node-level assignments for City H and City D ship with the network models,"
                 " which are available from the corresponding authors on reasonable request"
                 " (Data availability)."])

    readme(wb, "Source Data for Fig. 3 and Supplementary Figs. S1/S3 (zoning and sensor maps)",
           ["The maps are drawn from the network models plus the committed partition and"
            " sensor files; models for City H and City D are available on request, so this"
            " workbook carries the full assignment for the vendored EXA7 network and"
            " zone-size summaries for all five."],
           [["EXA7_node_to_zone", "Fig. 3a, Fig. S1a", "data/exa7/partitions.pkl +"
             " data/exa7/sensor_fingerprints.json", "[15].node_to_community, sensor_nodes",
             "data/exa7 setup (vendored)"],
            ["Zone_sizes_all_networks", "Fig. 3, Figs. S1/S3", "data/<network>/partitions.pkl +"
             " sensor_fingerprints.json", "[K].node_to_community", "data/*_setup.py"]])
    return save(wb, "Source_Data_Fig3_S1_S3.xlsx")


# ======================================================================
# workbook 6: Supplementary Data 1 (per-event records)
# ======================================================================
def build_sd1(dmap):
    lt = L("results_ltown.json")
    dg = L("results_city_d.json")
    wb = Workbook()

    lk = list(lt["rows"][0].keys())
    sheet(wb, "LTown_per_leak", lk, [[fmt(r.get(k)) for k in lk] for r in lt["rows"]],
          notes=["All 33 labelled BattLeDIM leaks diagnosed against the nominal model."])

    reg_fields = ["id", "type", "dma", "diameter_mm", "rate_Ls", "severity_tpd",
                  "max_abs_dp_clean_m", "max_abs_dp_m", "outcome", "true_zones",
                  "accepted_zone", "pred_zone", "top1_zone", "retrieval_top1",
                  "twinfit_top1", "correct", "existence", "margin", "top1_mahal", "best_alt"]
    sheet(wb, "CityD_register_per_event", reg_fields,
          [anon_row(r, dmap, reg_fields) for r in dg["rows"]],
          notes=["All 194 audited 2025 work orders; DMA names coded, leak types translated;"
                 " every numeric value as computed by eval/run_city_d.py."])

    bb = dg["by_severity_band"]
    bk = list(bb[0].keys())
    sheet(wb, "CityD_severity_bands", bk, [[fmt(b.get(k)) for k in bk] for b in bb])

    readme(wb, "Supplementary Data 1: per-event diagnostic records (L-Town and City D register)",
           ["The only two legs with individual real-world-labelled events; one row per event.",
            "City D DMA names are replaced by stable codes for utility confidentiality."],
           [["LTown_per_leak", "Fig. 5a/b, Supplementary Table S6", "artifacts/results_ltown.json",
             "rows", "eval/run_ltown.py"],
            ["CityD_register_per_event", "Fig. 5c, Supplementary Table S11",
             "artifacts/results_city_d.json", "rows", "eval/run_city_d.py"],
            ["CityD_severity_bands", "Supplementary Table S11", "artifacts/results_city_d.json",
             "by_severity_band", "eval/run_city_d.py"]])
    return save(wb, "Supplementary_Data_1.xlsx")


# ======================================================================
# workbook 7: Supplementary Data 2 (SI table backing, statistics legs)
# ======================================================================
def kv_sheet(wb, title, mapping, notes=()):
    sheet(wb, title, ["quantity", "value"], [[k, fmt(v)] for k, v in mapping], notes=notes)


def flat(prefix, obj):
    rows = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            rows += flat(prefix + "." + k if prefix else k, v)
    else:
        rows.append((prefix, obj))
    return rows


def build_sd2():
    st = L("results_stats.json")
    ab = L("results_ablation_components.json")
    cf = L("results_conformal.json")
    sv = L("results_sensitivity.json")
    lt = L("results_ltown.json")
    ky = L("results_ky4.json")
    hs = L("results_city_h_standard.json")
    ds = L("results_city_d_standard.json")
    d2 = L("results_city_d_v2_standard.json")
    pl = L("results_city_d_placement.json")
    p2 = L("results_city_d_placement_v2.json")
    dg = L("results_city_d.json")
    dv = L("results_city_d_severity.json")
    cl = L("results_city_d_ceiling_flow.json")
    ft = L("results_city_d_flowtier.json")
    dm = L("results_city_d_demand_sensitivity.json")
    mf = L("results_margin_fix_comparison.json")
    rt = L("results_runtime.json")
    wb = Workbook()

    ps = st["per_severity_leak_acc"]
    sheet(wb, "S1_severity_acc", ["injected discharge", "n", "accuracy (fraction)",
                                  "per-seed values"],
          [[k, ps[k]["n"], ps[k]["acc"], fmt(ps[k]["values"])] for k in ps])

    dcp = st["differential_confusion_pooled"]
    sheet(wb, "S2_confusion_pooled", ["true class \\ predicted"] + list(dcp["classes"]),
          [[c] + list(row) for c, row in zip(dcp["classes"], dcp["matrix"])])

    kv_sheet(wb, "S4_ablations",
             flat("no_differential", st["ablation_no_differential"])
             + flat("loosened_contract", st["ablation_loosened_contract"])
             + flat("components.full", ab["full"]) + flat("components.no_occam", ab["no_occam"])
             + flat("components.random_reveal", ab["random_reveal"])
             + [("components.no_active_leak_acc", ab["no_active_leak_acc"])])

    kv_sheet(wb, "S5_conformal_summary",
             flat("ece", cf["ece"]) + flat("precision_at_cov0.30", cf["precision_at_cov0.30"])
             + flat("split_conformal", cf["split_conformal"]) + flat("provenance", cf["provenance"]))
    cc = cf["calibrated_curve"]
    sheet(wb, "S5_calibrated_curve", list(cc[0].keys()),
          [[p[k] for k in cc[0].keys()] for p in cc])
    tc = cf["contract_curve"]
    sheet(wb, "S5_contract_curve", list(tc[0].keys()),
          [[p[k] for k in tc[0].keys()] for p in tc])

    fr = lt["risk_coverage"]["frontier"]
    sheet(wb, "S7_ltown_frontier", list(fr[0].keys()),
          [[p[k] for k in fr[0].keys()] for p in fr])

    for tag, art in (("KY4", ky), ("CityH", hs), ("CityD", ds), ("CityD_v2", d2)):
        h, r = metric_rows(art["multiseed_sigma0.05"], seeds_of(art["provenance"]))
        sheet(wb, "S8_{}".format(tag), h, r,
              notes=["Provenance: " + "; ".join("{} = {}".format(k, fmt(v))
                     for k, v in art["provenance"].items())])
    kv_sheet(wb, "S8_placement",
             flat("library_metrics_old", pl["library_metrics_old"])
             + flat("library_metrics_new", pl["library_metrics_new"])
             + flat("v2.detection_old", p2["detection_old"])
             + flat("v2.detection_new", p2["detection_new"])
             + flat("v2.adoption_decision", p2["adoption_decision"]))

    kv_sheet(wb, "S10_nominal", flat("nominal", sv["nominal"])
             + [("precision_min_across_sweeps", sv["precision_min_across_sweeps"]),
                ("precision_max_across_sweeps", sv["precision_max_across_sweeps"]),
                ("n_events_pooled", sv["provenance"].get("n_events_pooled"))])
    rows = []
    for pname, pts in sv["sweeps"].items():
        for p in pts:
            rows.append([pname] + [p.get(c) for c in ("value", "n_acted", "coverage",
                                                      "precision", "precision_per_seed_sd",
                                                      "n_seeds_with_actions")])
    sheet(wb, "S10_threshold_sweeps", ["parameter", "value", "n acted", "coverage",
                                       "precision", "precision per-seed sd",
                                       "n seeds with actions"], rows)

    bb = dg["by_severity_band"]
    sheet(wb, "S11_register_bands", list(bb[0].keys()),
          [[fmt(b.get(k)) for k in bb[0].keys()] for b in bb],
          notes=["Footer quantities of Table S11 come from results_city_d.json"
                 " detectability + summary and results_city_d_controls.json."])

    sw = dv["sweep"]
    sheet(wb, "S12_cityd_severity", list(sw[0].keys()),
          [[p[k] for k in sw[0].keys()] for p in sw])

    kv_sheet(wb, "S13_ceiling_flowtier",
             flat("ceiling", cl["ceiling"]) + [("n_events", cl["provenance"].get("n_events"))]
             + flat("adjacency_tolerance_check", ft["adjacency_tolerance_check"]))
    rows = [[k] + [fmt(v[c]) for c in ("sigma_f_Ls", "u", "n_events", "survey_dispatched",
                                       "zone_correct", "coverage", "precision_on_dispatched",
                                       "controls_false_alarm", "fa_rate")]
            for k, v in ft["results"].items()]
    sheet(wb, "S13_flowtier_settings", ["setting", "sigma_f (L/s)", "u", "n events",
                                        "survey dispatched", "zone correct", "coverage",
                                        "precision on dispatched", "controls false alarm",
                                        "false alarm rate"], rows)

    rows = []
    for ratio, blk in dm["ratios"].items():
        r = [ratio, blk.get("baseline_min_pressure_m"), blk.get("n_events")]
        r += [blk["ceiling"].get(c) for c in ("median_m", "p90_m", "n_ge_0.15", "n_ge_0.45")]
        r += [blk["flow"].get("massbal_median"), blk["flow"].get("detect_u3.4")]
        rows.append(r)
    sheet(wb, "S14_demand_sensitivity", ["demand ratio", "baseline min pressure (m)", "n events",
                                         "ceiling median (m)", "ceiling p90 (m)", "n >= 0.15 m",
                                         "n >= 0.45 m", "mass-balance median", "detect at u=3.4"],
          rows, notes=["The adopted ratio is marked in Supplementary Table S14."])

    rows = []
    for i, s in enumerate(mf["seeds"]):
        for tag in ("legacy", "fixed"):
            b = mf[tag][i]
            rows.append([s, tag] + [b.get(c) for c in sorted(b.keys())])
    cols = sorted(mf["legacy"][0].keys())
    sheet(wb, "MarginFix_per_seed", ["seed", "variant"] + cols, rows,
          notes=["Per-seed comparison of the legacy and corrected active-sensing"
                 " re-normalization (Methods); flips per seed: " + fmt(mf["flips"]) + "."])

    kv_sheet(wb, "Runtime", flat("", {k: v for k, v in rt.items() if k != "provenance"}))

    readme(wb, "Supplementary Data 2: numeric backing of the Supplementary Tables",
           ["One sheet per Supplementary Table (S1-S14 where the table is numeric),"
            " plus the margin-fix per-seed comparison and the runtime profile.",
            "Tables S3 and S6 are backed by Source Data Fig. 4 (sheet Fig4c_baselines)"
            " and Supplementary Data 1 (sheet LTown_per_leak) respectively.",
            "Table S9 and Table S15 are backed by Supplementary Data 3."],
           [["S1_severity_acc", "Table S1", "artifacts/results_stats.json", "per_severity_leak_acc",
             "eval/run_full_stats.py"],
            ["S2_confusion_pooled", "Table S2", "artifacts/results_stats.json",
             "differential_confusion_pooled", "eval/run_full_stats.py"],
            ["S4_ablations", "Table S4", "artifacts/results_stats.json +"
             " results_ablation_components.json", "ablation_*, full/no_occam/random_reveal",
             "eval/run_full_stats.py, eval/ablation_components.py"],
            ["S5_*", "Table S5", "artifacts/results_conformal.json",
             "ece, precision_at_cov0.30, split_conformal, curves", "eval/conformal.py"],
            ["S7_ltown_frontier", "Table S7", "artifacts/results_ltown.json",
             "risk_coverage.frontier", "eval/run_ltown.py"],
            ["S8_*", "Table S8", "results_{ky4,city_h_standard,city_d_standard,"
             "city_d_v2_standard,city_d_placement,city_d_placement_v2}.json",
             "multiseed_sigma0.05, provenance, placement metrics",
             "eval/run_ky4.py, eval/run_standard_net.py, eval/optimize_city_d_sensors*.py"],
            ["S10_*", "Table S10", "artifacts/results_sensitivity.json", "nominal, sweeps",
             "eval/contract_sensitivity.py"],
            ["S11_register_bands", "Table S11", "artifacts/results_city_d.json",
             "by_severity_band", "eval/run_city_d.py"],
            ["S12_cityd_severity", "Table S12", "artifacts/results_city_d_severity.json", "sweep",
             "eval/run_city_d_severity.py"],
            ["S13_*", "Table S13", "results_city_d_ceiling_flow.json + results_city_d_flowtier.json",
             "ceiling, results, adjacency_tolerance_check",
             "eval/run_city_d_ceiling_flow.py, eval/run_city_d_flow_tier.py"],
            ["S14_demand_sensitivity", "Table S14", "artifacts/results_city_d_demand_sensitivity.json",
             "ratios", "eval/run_city_d_demand_sensitivity.py"],
            ["MarginFix_per_seed", "Methods (margin correction)",
             "artifacts/results_margin_fix_comparison.json", "seeds, legacy, fixed, flips",
             "eval/compare_margin_fix.py"],
            ["Runtime", "Methods (runtime)", "artifacts/results_runtime.json", "all keys",
             "eval/runtime_profile.py"]])
    return save(wb, "Supplementary_Data_2.xlsx")


# ======================================================================
# workbook 8: Supplementary Data 3 (LLM audit experiments)
# ======================================================================
def build_sd3():
    ac = L("results_accountability.json")
    ex = L("results_llm_executor.json")
    ag = L("results_audit_generalization.json")
    wb = Workbook()

    kv_sheet(wb, "S9_audit_stress", flat("", {k: v for k, v in ac.items()}))
    kv_sheet(wb, "S9_dual_llm",
             flat("", {k: v for k, v in ex.items() if k != "sample_plans"}))
    plans = ex.get("sample_plans", [])
    if plans:
        pk = list(plans[0].keys())
        sheet(wb, "S9_sample_plans", ["plan"] + pk,
              [[i + 1] + [fmt(p.get(k)) for k in pk] for i, p in enumerate(plans)])

    kv_sheet(wb, "S15_provenance", flat("", ag["provenance"]))
    rows = [["RULES", "-", cls, blk["n"], blk["rejected"], ""]
            for cls, blk in ag["rules"].items()]
    for combo, classes in ag["llm"].items():
        stab = classes.get("_stability_out_of_taxonomy", "")
        for cls, blk in classes.items():
            if cls.startswith("_"):
                continue
            rows.append([combo.split("|")[0], combo.split("|")[1], cls,
                         blk["n"], blk["rejected"], stab])
    sheet(wb, "S15_summary", ["auditor", "prompt mode", "corruption class", "n", "rejected",
                              "stability (out-of-taxonomy)"], rows)

    rows = []
    for combo, classes in ag["llm"].items():
        model, mode = combo.split("|")
        for cls, blk in classes.items():
            if cls.startswith("_") or "decisions" not in blk:
                continue
            reasons = blk.get("reasons") or [""] * len(blk["decisions"])
            for i, (dec, why) in enumerate(zip(blk["decisions"], reasons)):
                rows.append([model, mode, cls, i + 1, "rejected" if dec else "passed", why])
    sheet(wb, "S15_per_case", ["auditor", "prompt mode", "corruption class", "case",
                               "verdict", "auditor reason (verbatim)"], rows,
          widths={6: 90})

    readme(wb, "Supplementary Data 3: language-model audit experiments",
           ["Per-case verdicts and verbatim reasons for the auditor stress test and the"
            " out-of-taxonomy generalization test; every underlying call is committed as"
            " a transcript in the code repository (artifacts/llm_transcripts_*.jsonl).",
            "Aggregates match Supplementary Tables S9 and S15."],
           [["S9_audit_stress", "Table S9", "artifacts/results_accountability.json", "all keys",
             "eval/run_accountability.py"],
            ["S9_dual_llm / S9_sample_plans", "Table S9", "artifacts/results_llm_executor.json",
             "all keys, sample_plans", "eval/run_llm_executor.py"],
            ["S15_provenance / S15_summary / S15_per_case", "Table S15",
             "artifacts/results_audit_generalization.json", "provenance, rules, llm.*",
             "eval/run_audit_generalization.py"]])
    return save(wb, "Supplementary_Data_3.xlsx")


# ======================================================================
# verification + manifest
# ======================================================================
def verify(paths):
    from openpyxl import load_workbook
    problems = []
    for p in paths:
        wb = load_workbook(p)
        for ws in wb.worksheets:
            if ws.max_row < 2 and ws.title != "README":
                problems.append("{}: sheet {} looks empty".format(os.path.basename(p), ws.title))
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str):
                        if "—" in c.value:
                            problems.append("{}:{} em-dash".format(os.path.basename(p), c.coordinate))
                        if CJK.search(c.value):
                            problems.append("{}:{}!{} CJK text: {!r}".format(
                                os.path.basename(p), ws.title, c.coordinate, c.value[:40]))
    return problems


def main():
    p1 = build_fig4()
    p2, dmap = build_fig5()
    p3 = build_table1()
    p4 = build_figs2()
    p5 = build_fig3_maps()
    p6 = build_sd1(dmap)
    p7 = build_sd2()
    p8 = build_sd3()
    paths = [p1, p2, p3, p4, p5, p6, p7, p8]

    with open(os.path.join(OUT, "PRIVATE_cityd_dma_map.json"), "w", encoding="utf-8") as fh:
        json.dump(dmap, fh, ensure_ascii=False, indent=2)

    problems = verify(paths)
    print("built {} workbooks in {}".format(len(paths), OUT))
    for p in paths:
        print("  ", os.path.basename(p), "{:.0f} KB".format(os.path.getsize(p) / 1024))
    if problems:
        print("PROBLEMS:")
        for q in problems:
            print("  ", q)
        sys.exit(1)
    print("verification clean: no empty sheets, no CJK, no em-dash in any submission cell")


if __name__ == "__main__":
    main()
